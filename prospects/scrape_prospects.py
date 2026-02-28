"""
AgentNexLiFy Prospect List Builder
Tries 3 approaches in order:
  1. googlesearch-python + website scraping
  2. Yelp search page scraping
  3. Final fallback: pre-built Google search URL spreadsheet
"""

import re
import time
import random
import shutil
from pathlib import Path
from urllib.parse import quote_plus, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ──
INDUSTRIES = [
    "plumber", "HVAC", "dentist", "real estate agent",
    "law firm", "med spa", "general contractor", "marketing agency",
]
CITIES = [
    "Austin TX", "Dallas TX", "Charlotte NC", "Tampa FL",
    "Denver CO", "Nashville TN", "Phoenix AZ", "Atlanta GA",
    "Portland OR", "Raleigh NC", "San Antonio TX", "Columbus OH",
    "Jacksonville FL", "Indianapolis IN", "Charleston SC",
]

OUTPUT_PATH = Path.home() / "agentnexlify" / "prospects" / "AgentNexLiFy-Prospect-List.xlsx"
COPY_PATH = Path("/mnt/user-data/outputs/AgentNexLiFy-Prospect-List-Populated.xlsx")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
PHONE_RE = re.compile(
    r"(?:\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}"
)

prospects = []  # list of dicts


def clean_phone(raw):
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) != 10:
        return None
    return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"


def clean_email(raw):
    raw = raw.lower().strip()
    if any(raw.endswith(f".{ext}") for ext in ["png", "jpg", "jpeg", "gif", "svg", "webp", "css", "js"]):
        return None
    if "@sentry" in raw or "@example" in raw or "noreply" in raw:
        return None
    return raw


def scrape_contact_from_url(url, timeout=5):
    """Visit a URL and extract phone/email via regex."""
    emails, phones = set(), set()
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        if r.status_code != 200:
            return emails, phones
        text = r.text[:200_000]  # cap at 200KB
        for m in EMAIL_RE.finditer(text):
            e = clean_email(m.group())
            if e:
                emails.add(e)
        for m in PHONE_RE.finditer(text):
            p = clean_phone(m.group())
            if p:
                phones.add(p)
    except Exception:
        pass
    return emails, phones


def extract_business_name(title, industry, city):
    """Try to clean a page title into a business name."""
    if not title:
        return None
    # Remove common suffixes
    for sep in [" | ", " - ", " :: ", " \u2013 ", " \u2014 "]:
        if sep in title:
            title = title.split(sep)[0]
    title = title.strip()
    if len(title) < 3 or len(title) > 80:
        return None
    return title


# ═══════════════════════════════════════════
# APPROACH 1: googlesearch-python
# ═══════════════════════════════════════════
def try_google_search():
    print("\n=== APPROACH 1: Google Search + Website Scraping ===")
    try:
        from googlesearch import search as gsearch
    except ImportError:
        print("  googlesearch-python not available, skipping.")
        return False

    found = 0
    blocked = False
    combos = [(ind, city) for ind in INDUSTRIES for city in CITIES]
    random.shuffle(combos)  # randomize to avoid pattern detection

    # Test with a single query first
    test_industry, test_city = combos[0]
    test_query = f"{test_industry} {test_city} phone email"
    print(f"  Testing with: \"{test_query}\"")
    try:
        results = list(gsearch(test_query, num_results=3, sleep_interval=3))
        if not results:
            print("  No results returned — likely blocked.")
            return False
        print(f"  Test OK — got {len(results)} results. Proceeding...")
    except Exception as e:
        print(f"  Google search blocked or errored: {e}")
        return False

    # Process all combos
    for i, (industry, city) in enumerate(combos):
        if blocked:
            break
        query = f"{industry} {city} phone email"
        print(f"  [{i+1}/{len(combos)}] \"{query}\"", end="", flush=True)

        try:
            results = list(gsearch(query, num_results=5, sleep_interval=0))
        except Exception as e:
            print(f" — BLOCKED: {e}")
            blocked = True
            break

        if not results:
            print(" — no results")
            time.sleep(random.uniform(3, 5))
            continue

        for url in results[:5]:
            domain = urlparse(url).netloc
            # Skip aggregator sites
            if any(skip in domain for skip in [
                "yelp.com", "yellowpages", "angi.com", "homeadvisor",
                "bbb.org", "facebook.com", "instagram.com", "twitter.com",
                "linkedin.com", "google.com", "youtube.com", "tiktok.com",
                "nextdoor.com", "thumbtack.com", "mapquest.com",
            ]):
                continue

            emails, phones = scrape_contact_from_url(url)
            if emails or phones:
                # Try to get a name from the page title
                try:
                    r = requests.get(url, headers=HEADERS, timeout=5)
                    soup = BeautifulSoup(r.text[:50000], "html.parser")
                    title = soup.title.string if soup.title else None
                    name = extract_business_name(title, industry, city)
                except Exception:
                    name = domain

                prospects.append({
                    "industry": industry,
                    "city": city,
                    "name": name or domain,
                    "website": url,
                    "email": "; ".join(sorted(emails)[:3]),
                    "phone": "; ".join(sorted(phones)[:2]),
                    "source": "Google Search",
                })
                found += 1

            time.sleep(random.uniform(1, 2))

        print(f" — {len([p for p in prospects if p['industry']==industry and p['city']==city])} found")
        time.sleep(random.uniform(3, 5))

    print(f"  Total from Google: {found} prospects")
    return found > 20  # consider success if we got a meaningful number


# ═══════════════════════════════════════════
# APPROACH 2: Yelp scraping
# ═══════════════════════════════════════════
def try_yelp_search():
    print("\n=== APPROACH 2: Yelp Search Scraping ===")
    found = 0
    blocked = False

    # Test first
    test_url = "https://www.yelp.com/search?find_desc=plumber&find_loc=Austin%2C+TX"
    print(f"  Testing Yelp access...")
    try:
        r = requests.get(test_url, headers=HEADERS, timeout=10)
        if r.status_code != 200:
            print(f"  Yelp returned {r.status_code} — likely blocked.")
            return False
        if "unusual traffic" in r.text.lower() or "captcha" in r.text.lower():
            print("  Yelp is showing captcha — blocked.")
            return False
        print(f"  Yelp accessible (status {r.status_code}). Proceeding...")
    except Exception as e:
        print(f"  Yelp not accessible: {e}")
        return False

    for industry in INDUSTRIES:
        if blocked:
            break
        for city in CITIES:
            if blocked:
                break
            yelp_url = f"https://www.yelp.com/search?find_desc={quote_plus(industry)}&find_loc={quote_plus(city)}"
            print(f"  {industry} in {city}", end="", flush=True)

            try:
                r = requests.get(yelp_url, headers=HEADERS, timeout=10)
                if r.status_code != 200:
                    print(f" — HTTP {r.status_code}")
                    if r.status_code == 429 or r.status_code == 503:
                        blocked = True
                    continue
                if "unusual traffic" in r.text.lower():
                    print(" — CAPTCHA blocked")
                    blocked = True
                    continue
            except Exception as e:
                print(f" — error: {e}")
                continue

            soup = BeautifulSoup(r.text, "html.parser")

            # Yelp's structure changes, try multiple selectors
            biz_links = soup.select('a[href*="/biz/"]')
            seen_names = set()
            page_found = 0

            for link in biz_links:
                href = link.get("href", "")
                if not href.startswith("/biz/") and not href.startswith("https://www.yelp.com/biz/"):
                    continue

                name_text = link.get_text(strip=True)
                if not name_text or len(name_text) < 3 or len(name_text) > 60:
                    continue
                # Skip navigation/filter links
                if name_text.lower() in ("write a review", "more", "yelp", "sign up", "log in"):
                    continue
                if name_text in seen_names:
                    continue
                seen_names.add(name_text)

                biz_url = href if href.startswith("http") else f"https://www.yelp.com{href}"

                # Try to get phone from Yelp biz page
                phone = ""
                try:
                    br = requests.get(biz_url, headers=HEADERS, timeout=5)
                    for m in PHONE_RE.finditer(br.text):
                        p = clean_phone(m.group())
                        if p:
                            phone = p
                            break
                except Exception:
                    pass

                prospects.append({
                    "industry": industry,
                    "city": city,
                    "name": name_text,
                    "website": biz_url,
                    "email": "",
                    "phone": phone,
                    "source": "Yelp",
                })
                found += 1
                page_found += 1

                if page_found >= 5:
                    break
                time.sleep(random.uniform(1, 2))

            print(f" — {page_found} found")
            time.sleep(random.uniform(3, 5))

    print(f"  Total from Yelp: {found} prospects")
    return found > 20


# ═══════════════════════════════════════════
# APPROACH 3 (FINAL FALLBACK): Search URL Spreadsheet
# ═══════════════════════════════════════════
def generate_search_url_sheet():
    print("\n=== APPROACH 3 (FINAL FALLBACK): Generating Search URL Spreadsheet ===")
    count = 0
    for industry in INDUSTRIES:
        for city in CITIES:
            query = f"{industry} {city} email phone"
            search_url = f"https://www.google.com/search?q={quote_plus(query)}"
            prospects.append({
                "industry": industry,
                "city": city,
                "name": "",
                "website": search_url,
                "email": "",
                "phone": "",
                "source": "Search URL (manual)",
            })
            count += 1

            # Also add a Maps search URL
            maps_query = f"{industry} near {city}"
            maps_url = f"https://www.google.com/maps/search/{quote_plus(maps_query)}"
            prospects.append({
                "industry": industry,
                "city": city,
                "name": "",
                "website": maps_url,
                "email": "",
                "phone": "",
                "source": "Maps URL (manual)",
            })
            count += 1

    print(f"  Generated {count} pre-built search URLs")
    return True


# ═══════════════════════════════════════════
# Spreadsheet Builder
# ═══════════════════════════════════════════
def build_spreadsheet():
    print("\n=== Building Spreadsheet ===")
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"

    # Styles
    header_font = Font(name="Inter", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    accent_fill = PatternFill("solid", fgColor="0a0a14")
    cell_font = Font(name="Inter", size=10, color="222222")
    link_font = Font(name="Inter", size=10, color="0066CC", underline="single")
    border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="EEEEEE"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    # Headers
    headers = ["#", "Industry", "City", "Business Name", "Website / Search URL", "Email", "Phone", "Source", "Notes"]
    col_widths = [5, 18, 18, 35, 50, 35, 20, 18, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(prospects) + 1}"

    # Data rows
    scraped = 0
    manual = 0
    for i, p in enumerate(prospects, 1):
        row = i + 1
        is_manual = "manual" in p["source"].lower()
        if is_manual:
            manual += 1
        else:
            scraped += 1

        fill = accent_fill if i % 2 == 0 else PatternFill()
        values = [
            i,
            p["industry"].title(),
            p["city"],
            p["name"],
            p["website"],
            p["email"],
            p["phone"],
            p["source"],
            "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = link_font if col == 5 and val else cell_font
            cell.fill = fill
            cell.border = border
            cell.alignment = wrap
            if col == 5 and val and val.startswith("http"):
                cell.hyperlink = val

        ws.row_dimensions[row].height = 22

    # Summary sheet
    summary = wb.create_sheet("Summary")
    summary_data = [
        ["AgentNexLiFy Prospect List — Summary"],
        [],
        ["Total Rows", len(prospects)],
        ["Scraped (with contact info)", scraped],
        ["Manual Search URLs", manual],
        [],
        ["Industries", len(INDUSTRIES)],
        ["Cities", len(CITIES)],
        ["Combinations", len(INDUSTRIES) * len(CITIES)],
        [],
        ["Industries Covered:"],
    ] + [[f"  {ind.title()}"] for ind in INDUSTRIES] + [
        [],
        ["Cities Covered:"],
    ] + [[f"  {c}"] for c in CITIES]

    for r, row_data in enumerate(summary_data, 1):
        for c, val in enumerate(row_data, 1):
            cell = summary.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(name="Inter", size=14, bold=True)
            elif c == 1 and isinstance(val, str) and not val.startswith(" "):
                cell.font = Font(name="Inter", size=11, bold=True)
            else:
                cell.font = Font(name="Inter", size=11)
    summary.column_dimensions["A"].width = 35
    summary.column_dimensions["B"].width = 15

    # Save
    wb.save(OUTPUT_PATH)
    print(f"  Saved to: {OUTPUT_PATH}")

    try:
        COPY_PATH.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(OUTPUT_PATH, COPY_PATH)
        print(f"  Copied to: {COPY_PATH}")
    except Exception as e:
        print(f"  Could not copy to {COPY_PATH}: {e}")

    return scraped, manual


# ═══════════════════════════════════════════
# Main
# ═══════════════════════════════════════════
def main():
    print("=" * 60)
    print("  AgentNexLiFy Prospect List Builder")
    print("=" * 60)
    print(f"  {len(INDUSTRIES)} industries x {len(CITIES)} cities = {len(INDUSTRIES)*len(CITIES)} combos")

    approach_used = "none"

    # Approach 1
    if try_google_search():
        approach_used = "Google Search"
    else:
        # Approach 2
        if try_yelp_search():
            approach_used = "Yelp"
        else:
            approach_used = "Search URLs (manual fallback)"

    # Always add search URLs for any missing industry+city combos
    existing = {(p["industry"], p["city"]) for p in prospects if p["source"] != "Search URL (manual)" and p["source"] != "Maps URL (manual)"}
    missing_combos = [(ind, city) for ind in INDUSTRIES for city in CITIES if (ind, city) not in existing]

    if missing_combos:
        print(f"\n  Filling {len(missing_combos)} missing combos with search URLs...")
        for industry, city in missing_combos:
            query = f"{industry} {city} email phone"
            search_url = f"https://www.google.com/search?q={quote_plus(query)}"
            prospects.append({
                "industry": industry,
                "city": city,
                "name": "",
                "website": search_url,
                "email": "",
                "phone": "",
                "source": "Search URL (manual)",
            })
            maps_query = f"{industry} near {city}"
            maps_url = f"https://www.google.com/maps/search/{quote_plus(maps_query)}"
            prospects.append({
                "industry": industry,
                "city": city,
                "name": "",
                "website": maps_url,
                "email": "",
                "phone": "",
                "source": "Maps URL (manual)",
            })

    # Build the spreadsheet
    scraped, manual = build_spreadsheet()

    # Summary
    print("\n" + "=" * 60)
    print("  SUMMARY")
    print("=" * 60)
    print(f"  Primary approach: {approach_used}")
    print(f"  Scraped prospects (with contact info): {scraped}")
    print(f"  Manual search URLs (for you to fill in): {manual}")
    print(f"  Total rows in spreadsheet: {len(prospects)}")
    print(f"  File: {OUTPUT_PATH}")
    print(f"  Copy: {COPY_PATH}")
    print("=" * 60)


if __name__ == "__main__":
    main()
