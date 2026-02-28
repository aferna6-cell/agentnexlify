#!/usr/bin/env python3
"""Clean up the prospect spreadsheet: remove bad entries, backfill with new searches."""

import csv
import os
import random
import re
import time
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests
from ddgs import DDGS

os.environ["PYTHONUNBUFFERED"] = "1"

OUTPUT_DIR = Path(__file__).parent
XLSX_PATH = OUTPUT_DIR / "AgentNexLiFy-Prospect-List.xlsx"
CSV_PATH = OUTPUT_DIR / "prospects-backup.csv"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

SKIP_EMAIL_DOMAINS = {
    "yelp.com", "yellowpages.com", "google.com", "facebook.com",
    "twitter.com", "instagram.com", "linkedin.com", "example.com",
    "sentry.io", "wixpress.com", "squarespace.com", "godaddy.com",
    "googleapis.com", "w3.org", "schema.org", "gstatic.com",
    "cloudflare.com", "wordpress.com", "wp.com", "gravatar.com",
}

DIRECTORY_DOMAINS = {
    "yelp.com", "yellowpages.com", "bbb.org", "google.com", "facebook.com",
    "linkedin.com", "instagram.com", "twitter.com", "tiktok.com",
    "homeadvisor.com", "angi.com", "thumbtack.com", "nextdoor.com",
    "mapquest.com", "superpages.com", "manta.com", "citysearch.com",
    "foursquare.com", "tripadvisor.com", "trustpilot.com",
    "glassdoor.com", "indeed.com", "buildzoom.com", "houzz.com",
    "avvo.com", "findlaw.com", "justia.com", "lawyers.com",
    "healthgrades.com", "zocdoc.com", "vitals.com", "realtor.com",
    "zillow.com", "redfin.com", "trulia.com", "homelight.com",
    "fastexpert.com", "realtrends.com", "usnews.com", "expertise.com",
    "classpass.com", "mindbody.com", "gympass.com",
    "wikipedia.org", "reddit.com", "youtube.com", "pinterest.com",
    "amazon.com", "apple.com",
}

http_session = requests.Session()
http_session.headers.update(HEADERS)

PHONE_RE = re.compile(r"(?:\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}")
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")

ddgs = DDGS()

# Bad name patterns
BAD_PATTERNS = [
    "real estate agents in", "homes for sale", "contact us", "contact",
    "providers", "best ", "top ", "12 best", "15 best", "10 best",
    "the best ", "best gym", "best dentist", "best plumber",
    "find", "top 15", "top 100", "top 10",
]


def is_bad_name(name: str) -> bool:
    n = name.lower().strip()
    if len(n) < 5:
        return True
    for pat in BAD_PATTERNS:
        if n.startswith(pat) or n == pat.strip():
            return True
    if n in ("home", "lohi", "hook", "about"):
        return True
    return False


def is_bad_email(email: str) -> bool:
    if not email:
        return False
    e = email.lower()
    if "myemail@" in e or "example" in e or "test@" in e:
        return True
    return False


def _is_business_domain(url: str) -> bool:
    try:
        domain = urlparse(url).netloc.lower().removeprefix("www.")
        return not any(domain.endswith(d) or domain == d for d in DIRECTORY_DOMAINS)
    except Exception:
        return False


def _extract_domain(url: str) -> str:
    try:
        return urlparse(url).netloc.lower().removeprefix("www.")
    except Exception:
        return ""


def _clean_business_name(title: str) -> str:
    for sep in [" | ", " - ", " – ", " — ", " :: ", " · "]:
        if sep in title:
            title = title.split(sep)[0]
    title = re.sub(
        r",?\s+(Austin|Dallas|Charlotte|Tampa|Denver|Nashville|Phoenix|Atlanta|Portland|Raleigh|San\s*Antonio|Columbus|Jacksonville|Indianapolis|Charleston)\b.*$",
        "", title, flags=re.IGNORECASE
    )
    return title.strip()


def _extract_phone(text: str) -> str:
    for m in PHONE_RE.findall(text):
        digits = re.sub(r"\D", "", m)
        if len(digits) == 11 and digits[0] == "1":
            digits = digits[1:]
        if len(digits) == 10:
            return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return ""


def _is_valid_email(email: str) -> bool:
    domain = email.lower().split("@")[1] if "@" in email else ""
    local = email.lower().split("@")[0] if "@" in email else ""
    if domain in SKIP_EMAIL_DOMAINS:
        return False
    if local in {"noreply", "no-reply", "donotreply", "mailer-daemon", "postmaster", "webmaster", "admin", "support", "abuse"}:
        return False
    if domain.endswith((".png", ".jpg", ".gif", ".svg", ".webp", ".css", ".js")):
        return False
    return True


def enrich(biz: dict) -> dict:
    url = biz.get("website", "")
    if not url or not url.startswith("http"):
        return biz
    pages = [url] + [url.rstrip("/") + s for s in ["/contact", "/contact-us", "/about"]]
    for page_url in pages:
        try:
            time.sleep(random.uniform(0.3, 0.8))
            resp = http_session.get(page_url, timeout=5, allow_redirects=True)
            if resp.status_code != 200:
                continue
            if "html" not in resp.headers.get("content-type", "").lower():
                continue
            text = resp.text
            if not biz.get("email"):
                for em in EMAIL_RE.findall(text):
                    if _is_valid_email(em):
                        biz["email"] = em
                        break
            if not biz.get("phone"):
                biz["phone"] = _extract_phone(text)
            if biz.get("email") and biz.get("phone"):
                break
        except Exception:
            continue
    return biz


def search_replacements(industry: str, city: str, state: str, needed: int, seen: set) -> list[dict]:
    """Search for replacement businesses."""
    queries_map = {
        "Real Estate": f"real estate agency {city} {state}",
        "Home Services": f"plumbing company {city} {state}",
        "Dental & Medical": f"dental clinic {city} {state}",
        "Law Firms": f"law office {city} {state}",
        "Med Spas & Wellness": f"medical spa {city} {state}",
        "Fitness & Gyms": f"fitness center {city} {state}",
        "Contractors": f"remodeling company {city} {state}",
        "Agencies & Consultants": f"marketing firm {city} {state}",
    }
    query = queries_map.get(industry, f"{industry} {city} {state}")
    time.sleep(random.uniform(1.5, 3.0))

    try:
        results = ddgs.text(query, max_results=12)
    except Exception as e:
        print(f"  Search error: {e}")
        return []

    replacements = []
    for r in results:
        url = r.get("href", "")
        title = r.get("title", "")
        body = r.get("body", "")
        if not _is_business_domain(url):
            continue
        name = _clean_business_name(title)
        if not name or len(name) < 5 or is_bad_name(name):
            continue
        domain = _extract_domain(url)
        name_key = name.lower().strip()
        if domain in seen or name_key in seen:
            continue
        seen.add(domain)
        seen.add(name_key)

        phone = _extract_phone(title) or _extract_phone(body)
        biz = {
            "name": name, "phone": phone, "website": url,
            "city": city, "state": state, "source": "DuckDuckGo",
            "contact_name": "", "email": "",
        }
        print(f"    Enriching replacement: {name[:40]}...")
        biz = enrich(biz)
        if biz.get("email"):
            print(f"      Email: {biz['email']}")
        if biz.get("phone"):
            print(f"      Phone: {biz['phone']}")

        replacements.append(biz)
        if len(replacements) >= needed:
            break

    return replacements


def main():
    print("Cleaning up prospect spreadsheet...")

    # Load all data from CSV (easier to manipulate)
    rows = []
    with open(CSV_PATH, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    print(f"Loaded {len(rows)} rows from CSV")

    # Collect all seen domains/names
    seen = set()
    for row in rows:
        domain = _extract_domain(row.get("website", ""))
        if domain:
            seen.add(domain)
        seen.add(row["name"].lower().strip())

    # Identify bad rows
    bad_indices = []
    for i, row in enumerate(rows):
        if is_bad_name(row["name"]):
            bad_indices.append(i)
        elif is_bad_email(row.get("email", "")):
            row["email"] = ""  # Just clear the email, keep the row

    print(f"Found {len(bad_indices)} bad entries to replace")

    # Group bad entries by industry to replace efficiently
    by_industry = {}
    for i in bad_indices:
        ind = rows[i]["industry"]
        by_industry.setdefault(ind, []).append(i)

    # Search for replacements
    cities = [
        ("Austin", "TX"), ("Dallas", "TX"), ("Charlotte", "NC"),
        ("Tampa", "FL"), ("Denver", "CO"), ("Nashville", "TN"),
        ("Phoenix", "AZ"), ("Atlanta", "GA"), ("Portland", "OR"),
        ("Raleigh", "NC"), ("San Antonio", "TX"), ("Columbus", "OH"),
        ("Jacksonville", "FL"), ("Indianapolis", "IN"), ("Charleston", "SC"),
    ]

    for industry, indices in by_industry.items():
        needed = len(indices)
        print(f"\n  Replacing {needed} entries for {industry}")
        city, state = random.choice(cities)
        replacements = search_replacements(industry, city, state, needed, seen)

        for j, idx in enumerate(indices):
            if j < len(replacements):
                r = replacements[j]
                rows[idx] = {
                    "name": r["name"],
                    "contact_name": r.get("contact_name", ""),
                    "email": r.get("email", ""),
                    "phone": r.get("phone", ""),
                    "industry": industry,
                    "city": r.get("city", city),
                    "state": r.get("state", state),
                    "website": r.get("website", ""),
                    "address": r.get("address", ""),
                    "source": r.get("source", "DuckDuckGo"),
                    "status": "Not Contacted",
                }
                print(f"    Replaced with: {r['name']}")
            else:
                # Can't find replacement, just remove
                rows[idx] = None

    # Remove None entries
    rows = [r for r in rows if r is not None]
    print(f"\nFinal row count: {len(rows)}")

    # Rewrite spreadsheet
    from populate_prospects import COLUMNS, HEADER_FILL, HEADER_FONT, CELL_FONT, THIN_BORDER, _style_header

    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "All Prospects"
    _style_header(ws_all)

    # Group by industry
    by_ind = {}
    for r in rows:
        by_ind.setdefault(r["industry"], []).append(r)

    row_num = 2
    for industry in ["Real Estate", "Home Services", "Dental & Medical", "Law Firms",
                      "Med Spas & Wellness", "Fitness & Gyms", "Contractors", "Agencies & Consultants"]:
        bizlist = by_ind.get(industry, [])
        for biz in bizlist:
            values = [
                biz.get("name", ""), biz.get("contact_name", ""),
                biz.get("email", ""), biz.get("phone", ""),
                industry, f"{biz.get('city', '')}, {biz.get('state', '')}",
                biz.get("website", ""), biz.get("source", ""),
                "Not Contacted", "",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws_all.cell(row=row_num, column=col_idx, value=val)
                cell.font = CELL_FONT
                cell.border = THIN_BORDER
            row_num += 1

    # Per-industry sheets
    for industry, bizlist in by_ind.items():
        ws = wb.create_sheet(title=industry[:31])
        _style_header(ws)
        for i, biz in enumerate(bizlist, 2):
            values = [
                biz.get("name", ""), biz.get("contact_name", ""),
                biz.get("email", ""), biz.get("phone", ""),
                industry, f"{biz.get('city', '')}, {biz.get('state', '')}",
                biz.get("website", ""), biz.get("source", ""),
                "Not Contacted", "",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=i, column=col_idx, value=val)
                cell.font = CELL_FONT
                cell.border = THIN_BORDER

    # Summary
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 15
    ws_summary.cell(row=1, column=1, value="AgentNexLiFy Prospect Summary").font = Font(
        name="Calibri", bold=True, size=14
    )
    ws_summary.cell(row=3, column=1, value="Industry").font = Font(bold=True)
    ws_summary.cell(row=3, column=2, value="Count").font = Font(bold=True)
    ws_summary.cell(row=3, column=3, value="With Email").font = Font(bold=True)

    r = 4
    total = total_email = 0
    for industry, bizlist in by_ind.items():
        count = len(bizlist)
        email_count = sum(1 for b in bizlist if b.get("email"))
        ws_summary.cell(row=r, column=1, value=industry)
        ws_summary.cell(row=r, column=2, value=count)
        ws_summary.cell(row=r, column=3, value=email_count)
        total += count
        total_email += email_count
        r += 1
    r += 1
    ws_summary.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws_summary.cell(row=r, column=2, value=total).font = Font(bold=True)
    ws_summary.cell(row=r, column=3, value=total_email).font = Font(bold=True)

    wb.save(XLSX_PATH)
    print(f"Saved: {XLSX_PATH}")

    # Update CSV
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "name", "contact_name", "email", "phone", "industry",
            "city", "state", "website", "address", "source", "status",
        ])
        writer.writeheader()
        for biz in rows:
            writer.writerow({k: biz.get(k, "") for k in writer.fieldnames})
    print(f"Saved: {CSV_PATH}")

    # Final summary
    total = len(rows)
    emails = sum(1 for r in rows if r.get("email"))
    phones = sum(1 for r in rows if r.get("phone"))
    print(f"\n=== Final Summary ===")
    print(f"Total: {total}")
    print(f"With email: {emails}")
    print(f"With phone: {phones}")


if __name__ == "__main__":
    main()
