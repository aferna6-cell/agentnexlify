#!/usr/bin/env python3
"""
AgentNexLiFy Prospect List Populator

Uses DuckDuckGo search to find real business prospects, extracts contact info
from search results and business websites, and writes results to an Excel
spreadsheet and CSV backup.

Target: 200 real businesses across 8 industries and 15 cities.
"""

import csv
import os
import random
import re
import sys
import time
from pathlib import Path
from urllib.parse import urljoin, urlparse

# Force unbuffered output
os.environ["PYTHONUNBUFFERED"] = "1"

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests
from bs4 import BeautifulSoup
from ddgs import DDGS

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

OUTPUT_DIR = Path(__file__).parent
XLSX_PATH = OUTPUT_DIR / "AgentNexLiFy-Prospect-List.xlsx"
CSV_PATH = OUTPUT_DIR / "prospects-backup.csv"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

SEARCH_DELAY = (1.5, 3.0)  # delay between DuckDuckGo searches
EMAIL_TIMEOUT = 5
MAX_PER_INDUSTRY = 25
TARGET_TOTAL = 200

# Skip platform/junk email domains
SKIP_EMAIL_DOMAINS = {
    "yelp.com", "yellowpages.com", "google.com", "facebook.com",
    "twitter.com", "instagram.com", "linkedin.com", "example.com",
    "sentry.io", "wixpress.com", "squarespace.com", "godaddy.com",
    "googleapis.com", "w3.org", "schema.org", "gstatic.com",
    "cloudflare.com", "wordpress.com", "wp.com", "gravatar.com",
}
SKIP_EMAIL_PREFIXES = {
    "noreply", "no-reply", "donotreply", "mailer-daemon", "postmaster",
    "webmaster", "admin", "support", "abuse",
}

# Domains that are directories/aggregators, not actual businesses
DIRECTORY_DOMAINS = {
    "yelp.com", "yellowpages.com", "bbb.org", "google.com", "facebook.com",
    "linkedin.com", "instagram.com", "twitter.com", "tiktok.com",
    "homeadvisor.com", "angi.com", "thumbtack.com", "nextdoor.com",
    "mapquest.com", "superpages.com", "manta.com", "citysearch.com",
    "foursquare.com", "tripadvisor.com", "trustpilot.com",
    "glassdoor.com", "indeed.com", "buildzoom.com", "houzz.com",
    "avvo.com", "findlaw.com", "justia.com", "lawyers.com",
    "healthgrades.com", "zocdoc.com", "vitals.com", "realtor.com",
    "zillow.com", "redfin.com", "trulia.com", "homelight.com",
    "fastexpert.com", "realtrends.com", "usnews.com",
    "classpass.com", "mindbody.com", "gympass.com",
    "wikipedia.org", "reddit.com", "youtube.com", "pinterest.com",
    "amazon.com", "apple.com",
}

# ---------------------------------------------------------------------------
# Industries and cities
# ---------------------------------------------------------------------------

INDUSTRIES = {
    "Real Estate": [
        "real estate agent {city} {state} phone",
        "realtor {city} {state} contact",
        "real estate broker {city} {state}",
    ],
    "Home Services": [
        "plumber {city} {state} phone",
        "hvac company {city} {state} contact",
        "plumbing company {city} {state}",
    ],
    "Dental & Medical": [
        "dentist {city} {state} phone",
        "dental office {city} {state} contact",
        "family dentist {city} {state}",
    ],
    "Law Firms": [
        "law firm {city} {state} phone",
        "attorney {city} {state} contact",
        "lawyer {city} {state}",
    ],
    "Med Spas & Wellness": [
        "med spa {city} {state} phone",
        "medical spa {city} {state} contact",
        "aesthetics clinic {city} {state}",
    ],
    "Fitness & Gyms": [
        "gym {city} {state} phone",
        "fitness studio {city} {state} contact",
        "personal trainer {city} {state}",
    ],
    "Contractors": [
        "general contractor {city} {state} phone",
        "home remodeling {city} {state} contact",
        "remodeling contractor {city} {state}",
    ],
    "Agencies & Consultants": [
        "marketing agency {city} {state} phone",
        "digital marketing agency {city} {state}",
        "business consultant {city} {state} contact",
    ],
}

CITIES = [
    ("Austin", "TX"), ("Dallas", "TX"), ("Charlotte", "NC"),
    ("Tampa", "FL"), ("Denver", "CO"), ("Nashville", "TN"),
    ("Phoenix", "AZ"), ("Atlanta", "GA"), ("Portland", "OR"),
    ("Raleigh", "NC"), ("San Antonio", "TX"), ("Columbus", "OH"),
    ("Jacksonville", "FL"), ("Indianapolis", "IN"), ("Charleston", "SC"),
]

# ---------------------------------------------------------------------------
# Networking
# ---------------------------------------------------------------------------

http_session = requests.Session()
http_session.headers.update(HEADERS)

PHONE_RE = re.compile(
    r"(?:\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}"
)
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")


def _is_business_domain(url: str) -> bool:
    """Check if URL belongs to an actual business (not a directory)."""
    try:
        domain = urlparse(url).netloc.lower()
        # Strip www.
        if domain.startswith("www."):
            domain = domain[4:]
        return not any(domain.endswith(d) or domain == d for d in DIRECTORY_DOMAINS)
    except Exception:
        return False


def _extract_domain(url: str) -> str:
    """Get clean domain from URL."""
    try:
        domain = urlparse(url).netloc.lower()
        if domain.startswith("www."):
            domain = domain[4:]
        return domain
    except Exception:
        return ""


def _clean_business_name(title: str) -> str:
    """Extract business name from search result title."""
    # Remove common suffixes
    for sep in [" | ", " - ", " – ", " — ", " :: ", " · "]:
        if sep in title:
            title = title.split(sep)[0]
    # Remove trailing city/state info
    title = re.sub(r",?\s+(Austin|Dallas|Charlotte|Tampa|Denver|Nashville|Phoenix|Atlanta|Portland|Raleigh|San\s*Antonio|Columbus|Jacksonville|Indianapolis|Charleston)\b.*$", "", title, flags=re.IGNORECASE)
    return title.strip()


def _extract_phone_from_text(text: str) -> str:
    """Extract phone number from text."""
    matches = PHONE_RE.findall(text)
    for m in matches:
        digits = re.sub(r"\D", "", m)
        if len(digits) == 10 or (len(digits) == 11 and digits[0] == "1"):
            # Format nicely
            if len(digits) == 11:
                digits = digits[1:]
            return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return ""


# ---------------------------------------------------------------------------
# DuckDuckGo search
# ---------------------------------------------------------------------------

ddgs = DDGS()


def search_businesses(query: str, max_results: int = 10) -> list[dict]:
    """Search DuckDuckGo for business listings."""
    time.sleep(random.uniform(*SEARCH_DELAY))
    try:
        results = ddgs.text(query, max_results=max_results)
        return results
    except Exception as e:
        print(f"  [SEARCH ERROR] {e}")
        return []


def parse_search_results(
    results: list[dict], city: str, state: str
) -> list[dict]:
    """Parse DuckDuckGo results into business records."""
    businesses = []

    for r in results:
        url = r.get("href", "")
        title = r.get("title", "")
        body = r.get("body", "")

        # Skip directory/aggregator sites
        if not _is_business_domain(url):
            continue

        name = _clean_business_name(title)
        if not name or len(name) < 3:
            continue

        # Extract phone from title or body
        phone = _extract_phone_from_text(title) or _extract_phone_from_text(body)

        businesses.append({
            "name": name,
            "phone": phone,
            "website": url,
            "address": "",
            "city": city,
            "state": state,
            "source": "DuckDuckGo",
            "contact_name": "",
            "email": "",
        })

    return businesses


# ---------------------------------------------------------------------------
# Email + phone finder (from business websites)
# ---------------------------------------------------------------------------

def _is_valid_business_email(email: str) -> bool:
    """Filter out platform/junk emails."""
    email_lower = email.lower()
    domain = email_lower.split("@")[1] if "@" in email_lower else ""
    local = email_lower.split("@")[0] if "@" in email_lower else ""

    if domain in SKIP_EMAIL_DOMAINS:
        return False
    if local in SKIP_EMAIL_PREFIXES:
        return False
    if domain.endswith((".png", ".jpg", ".gif", ".svg", ".webp", ".css", ".js")):
        return False
    return True


def enrich_from_website(biz: dict) -> dict:
    """Visit business website to find email and phone if missing."""
    url = biz.get("website", "")
    if not url or not url.startswith("http"):
        return biz

    pages_to_try = [url]
    base = url.rstrip("/")
    for suffix in ["/contact", "/contact-us", "/about", "/about-us"]:
        pages_to_try.append(base + suffix)

    for page_url in pages_to_try:
        try:
            time.sleep(random.uniform(0.3, 1.0))
            resp = http_session.get(page_url, timeout=EMAIL_TIMEOUT, allow_redirects=True)
            if resp.status_code != 200:
                continue
            ct = resp.headers.get("content-type", "")
            if "html" not in ct.lower():
                continue

            text = resp.text

            # Find email
            if not biz.get("email"):
                emails = EMAIL_RE.findall(text)
                for em in emails:
                    if _is_valid_business_email(em):
                        biz["email"] = em
                        break

            # Find phone if missing
            if not biz.get("phone"):
                phone = _extract_phone_from_text(text)
                if phone:
                    biz["phone"] = phone

            # If we have both, stop
            if biz.get("email") and biz.get("phone"):
                break

        except requests.RequestException:
            continue

    return biz


# ---------------------------------------------------------------------------
# Spreadsheet writer
# ---------------------------------------------------------------------------

COLUMNS = [
    "Business Name", "Contact Name", "Email", "Phone",
    "Industry", "City/State", "Website", "Source", "Status", "Notes",
]

HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _style_header(ws):
    for col_idx, title in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    ws.freeze_panes = "A2"
    widths = [30, 20, 30, 16, 22, 20, 40, 14, 16, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_row(ws, row_num: int, biz: dict, industry: str):
    values = [
        biz.get("name", ""),
        biz.get("contact_name", ""),
        biz.get("email", ""),
        biz.get("phone", ""),
        industry,
        f"{biz.get('city', '')}, {biz.get('state', '')}",
        biz.get("website", ""),
        biz.get("source", ""),
        "Not Contacted",
        "",
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = CELL_FONT
        cell.border = THIN_BORDER


def write_spreadsheet(all_businesses: dict[str, list[dict]]):
    wb = openpyxl.Workbook()

    # "All Prospects" sheet
    ws_all = wb.active
    ws_all.title = "All Prospects"
    _style_header(ws_all)

    row = 2
    flat_list = []
    for industry, bizlist in all_businesses.items():
        for biz in bizlist:
            _write_row(ws_all, row, biz, industry)
            flat_list.append({**biz, "industry": industry})
            row += 1

    # Per-industry sheets
    for industry, bizlist in all_businesses.items():
        ws = wb.create_sheet(title=industry[:31])
        _style_header(ws)
        for i, biz in enumerate(bizlist, 2):
            _write_row(ws, i, biz, industry)

    # Summary sheet
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 15

    ws_summary.cell(row=1, column=1, value="AgentNexLiFy Prospect Summary").font = Font(
        name="Calibri", bold=True, size=14
    )
    ws_summary.cell(row=3, column=1, value="Industry").font = Font(bold=True)
    ws_summary.cell(row=3, column=2, value="Count").font = Font(bold=True)
    ws_summary.cell(row=3, column=3, value="With Email").font = Font(bold=True)

    r = 4
    total = total_email = 0
    for industry, bizlist in all_businesses.items():
        count = len(bizlist)
        email_count = sum(1 for b in bizlist if b.get("email"))
        ws_summary.cell(row=r, column=1, value=industry)
        ws_summary.cell(row=r, column=2, value=count)
        ws_summary.cell(row=r, column=3, value=email_count)
        total += count
        total_email += email_count
        r += 1

    r += 1
    ws_summary.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws_summary.cell(row=r, column=2, value=total).font = Font(bold=True)
    ws_summary.cell(row=r, column=3, value=total_email).font = Font(bold=True)

    wb.save(XLSX_PATH)
    print(f"\nSaved spreadsheet: {XLSX_PATH}")

    # CSV backup
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "name", "contact_name", "email", "phone", "industry",
            "city", "state", "website", "address", "source", "status",
        ])
        writer.writeheader()
        for biz in flat_list:
            writer.writerow({
                "name": biz.get("name", ""),
                "contact_name": biz.get("contact_name", ""),
                "email": biz.get("email", ""),
                "phone": biz.get("phone", ""),
                "industry": biz.get("industry", ""),
                "city": biz.get("city", ""),
                "state": biz.get("state", ""),
                "website": biz.get("website", ""),
                "address": biz.get("address", ""),
                "source": biz.get("source", ""),
                "status": "Not Contacted",
            })
    print(f"Saved CSV backup: {CSV_PATH}")


# ---------------------------------------------------------------------------
# Main orchestration
# ---------------------------------------------------------------------------

def gather_prospects() -> dict[str, list[dict]]:
    """Main search loop across industries and cities."""
    all_businesses: dict[str, list[dict]] = {ind: [] for ind in INDUSTRIES}
    seen_keys: set[str] = set()  # dedup by domain or name

    # Pre-shuffle cities per industry for geographic diversity
    city_lists = {ind: list(CITIES) for ind in INDUSTRIES}
    for ind in city_lists:
        random.shuffle(city_lists[ind])

    for industry, queries in INDUSTRIES.items():
        print(f"\n{'='*60}")
        print(f"Industry: {industry}")
        print(f"{'='*60}")

        collected = 0
        cities = city_lists[industry]
        query_idx = 0

        for city, state in cities:
            if collected >= MAX_PER_INDUSTRY:
                break

            query_template = queries[query_idx % len(queries)]
            query = query_template.format(city=city, state=state)
            query_idx += 1

            print(f"\n  Searching: {query}")
            results = search_businesses(query, max_results=12)
            print(f"  Got {len(results)} raw results")

            parsed = parse_search_results(results, city, state)
            print(f"  Parsed {len(parsed)} business results")

            for biz in parsed:
                if collected >= MAX_PER_INDUSTRY:
                    break

                # Dedup by domain or name
                domain = _extract_domain(biz["website"])
                name_key = biz["name"].lower().strip()
                if domain and domain in seen_keys:
                    continue
                if name_key in seen_keys:
                    continue
                if domain:
                    seen_keys.add(domain)
                seen_keys.add(name_key)

                # Enrich: try to find email + phone from website
                print(f"    Enriching: {biz['name'][:40]}...")
                biz = enrich_from_website(biz)
                if biz.get("email"):
                    print(f"      Email: {biz['email']}")
                if biz.get("phone"):
                    print(f"      Phone: {biz['phone']}")

                all_businesses[industry].append(biz)
                collected += 1

            print(f"  Progress: {collected}/{MAX_PER_INDUSTRY} {industry}")

        print(f"\n  DONE: {collected} {industry} prospects")

    return all_businesses


def print_summary(all_businesses: dict[str, list[dict]]):
    total = total_email = 0
    print(f"\n{'='*50}")
    print("=== AgentNexLiFy Prospect List Summary ===")
    print(f"{'='*50}")

    for industry, bizlist in all_businesses.items():
        total += len(bizlist)
        total_email += sum(1 for b in bizlist if b.get("email"))

    print(f"Total businesses found: {total}")
    print(f"With email: {total_email}")
    print(f"Without email: {total - total_email}")
    print()
    print("By industry:")
    for industry, bizlist in all_businesses.items():
        email_c = sum(1 for b in bizlist if b.get("email"))
        phone_c = sum(1 for b in bizlist if b.get("phone"))
        print(f"  {industry}: {len(bizlist)} ({email_c} emails, {phone_c} phones)")
    print()


def main():
    print("AgentNexLiFy Prospect List Populator")
    print("=" * 40)
    print(f"Target: {TARGET_TOTAL} businesses across {len(INDUSTRIES)} industries")
    print(f"Searching {len(CITIES)} cities\n")

    all_businesses = gather_prospects()

    write_spreadsheet(all_businesses)
    print_summary(all_businesses)

    # Copy to outputs if available
    outputs_dir = Path("/mnt/user-data/outputs")
    if outputs_dir.exists():
        import shutil
        dest = outputs_dir / "AgentNexLiFy-Prospect-List-Populated.xlsx"
        shutil.copy2(XLSX_PATH, dest)
        print(f"Copied to: {dest}")

    print("Done!")


if __name__ == "__main__":
    main()
